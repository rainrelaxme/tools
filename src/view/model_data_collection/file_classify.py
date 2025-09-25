import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def classify_images_by_defect(csv_file_path, source_directory):
    """
    根据CSV文件中的缺陷标记分类图片和JSON文件，并回填文件路径到Excel
    """

    # 读取CSV文件
    try:
        df = pd.read_excel(csv_file_path, engine='openpyxl')
        print(f"成功读取CSV文件，共 {len(df)} 行数据")
    except Exception as e:
        print(f"读取CSV文件失败: {e}")
        return

    # 检查必要的列是否存在
    required_columns = ['图片名称', '缺陷']
    for col in required_columns:
        if col not in df.columns:
            print(f"CSV文件中缺少必要的列: {col}")
            return

    # 处理数据
    processed_count = 0
    ok_count = 0
    ng_count = 0

    # 获取所有有效的图片文件名（去除空行）
    valid_rows = df[df['图片名称'].notna() & (df['图片名称'] != '')]

    # 用于存储文件路径信息
    file_paths = []

    for index, row in df.iterrows():
        index = index
        image_filename = str(row['图片名称'])
        defect = row['缺陷']

        # 确定目标文件夹
        # if defect == '无':
        #     # target_dir = ok_dir
        #     ok_count += 1
        #     status = 'OK'
        # else:
        #     # target_dir = ng_dir
        #     ng_count += 1
        #     status = 'NG'
        if image_filename:
            new_filename = image_filename[12:]
        else:
            new_filename = image_filename
        status = search_files_in_target(source_directory, new_filename)

        file_paths.append(status)

        processed_count += 1

    # 回填文件路径到Excel
    try:
        # 使用openpyxl来处理Excel文件，保留格式
        wb = load_workbook(csv_file_path)
        ws = wb.active

        # 添加表头（如果不存在）
        if ws.max_column < len(df.columns) + 1:
            header_cell = ws.cell(row=1, column=len(df.columns) + 1)
            header_cell.value = "自动标注结果"
            # 设置表头样式
            header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 填充文件路径数据
        for i, file_path in enumerate(file_paths, start=2):  # 从第2行开始（跳过表头）
            if i <= ws.max_row:

                ws.cell(row=i, column=len(df.columns) + 1).value = file_path

        # 保存Excel文件
        output_excel_path = csv_file_path.replace('.xlsx', '_with_status.xlsx')
        wb.save(output_excel_path)
        print(f"\n文件路径已回填到: {output_excel_path}")

    except Exception as e:
        print(f"回填文件路径到Excel时出错: {e}")
        # 如果Excel操作失败，尝试在DataFrame中添加列并保存为新的CSV
        try:
            df['文件路径'] = file_paths + [''] * (len(df) - len(file_paths))
            output_csv_path = csv_file_path.replace('.csv', '_with_paths.csv')
            df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
            print(f"文件路径已回填到: {output_csv_path}")
        except Exception as e2:
            print(f"保存CSV文件时也出错: {e2}")

    print(f"\n处理完成！")
    print(f"总共处理: {processed_count} 个文件")
    # print(f"OK文件夹: {ok_count} 个文件")
    # print(f"NG文件夹: {ng_count} 个文件")


def search_files_in_target(destination_directory, search_filename):
    """
    在目标文件夹中搜索文件
    """
    found_files = []
    status = ''
    # 搜索OK文件夹
    ok_dir = os.path.join(destination_directory, 'OK')
    if os.path.exists(ok_dir):
        for root, dirs, files in os.walk(ok_dir):
            if search_filename in files:
                status = 'OK'
                break

    # 搜索NG文件夹
    ng_dir = os.path.join(destination_directory, 'NG')
    if os.path.exists(ng_dir):
        for root, dirs, files in os.walk(ng_dir):
            if search_filename in files:
                status = 'NG'
                break

    # return found_files
    return status


def main():

    # 分类文件模式
    csv_file_path = input("请输入csv文件路径: ").strip()
    source_directory = input("请输入自动标注分类后文件目录路径: ").strip()

    df = pd.read_csv(csv_file_path, encoding='utf-8')
    excel_file_path = csv_file_path.replace('.csv', '.xlsx')
    df.to_excel(excel_file_path, index=False)

    # 执行分类
    classify_images_by_defect(excel_file_path, source_directory)


# if __name__ == "__main__":
#     main()