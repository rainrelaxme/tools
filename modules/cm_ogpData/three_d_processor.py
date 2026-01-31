#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Project : tools
@File    : three_d_processor.py
@Author  : Shawn
@Date    : 2026/1/8 9:45
@Info    : 三次元数据汇总程序
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ThreeDProcessor:
    def __init__(self):
        pass

    def merge_three_d_files(self, file_paths, output_dir, processing_flag, create_subfolder=True, layout_mode="horizontal"):
        """合并多个三次元Excel文件"""
        try:
            # 存储所有文件的数据
            all_data = {}  # {(尺寸, 轴): {nominal, upper_tol, lower_tol, measurements: {模序号: {穴序号: 值}}}}
            product_names = []  # 存储所有文件的产品名

            # 读取所有文件
            for file_idx, file_path in enumerate(file_paths, 1):
                if not processing_flag:
                    return False, "处理已停止", None

                # 解析文件名
                file_info = self.parse_filename(os.path.basename(file_path))
                if not file_info:
                    continue
                
                product_name, mold_num, cavity_range = file_info
                product_names.append(product_name)

                # 读取Excel文件（尝试读取所有工作表）
                try:
                    # 先尝试读取第一个工作表
                    df = pd.read_excel(file_path, header=None, sheet_name=0)
                except Exception as e:
                    continue

                # 检查文件是否有数据
                if df.empty:
                    continue

                # 查找数据开始行（跳过标题行）
                data_start_row = self.find_data_start_row(df)

                if data_start_row == -1:
                    continue

                # 解析列索引
                col_indices = self.detect_column_indices(df, data_start_row)

                if not col_indices:
                    continue

                # 提取数据
                # 检查文件列数，确保能够访问所需的列（至少需要8列）
                max_cols = max([len(row) for row in df.values] + [len(df.columns)]) if len(df) > 0 else len(df.columns)
                if max_cols < 8:
                    continue
                
                # 解析穴数范围
                if cavity_range is None:
                    continue
                
                start_cavity, end_cavity = cavity_range
                
                # 为每个尺寸维护独立的穴数索引
                size_cavity_indices = {}
                
                for row_idx in range(data_start_row, len(df)):
                    row = df.iloc[row_idx]
                    
                    # 检查行是否有足够的列
                    if len(row) < 8:
                        continue
                    
                    # 只处理第1列（尺寸）和第4列（轴）都有数据的行
                    size = self.get_cell_value(row, col_indices['size'])         # 第1列：尺寸
                    axis = self.get_cell_value(row, col_indices['axis'])         # 第4列：轴

                    # 跳过无效行：第1列或第4列为空的行
                    if size is None or pd.isna(size) or str(size).strip() == '':
                        continue
                    if axis is None or pd.isna(axis) or str(axis).strip() == '':
                        continue

                    # 提取其他列的数据
                    nominal = self.get_cell_value(row, col_indices['nominal'])         # 第5列：标准值
                    upper_tol = self.get_cell_value(row, col_indices['upper_tol'])    # 第7列：上公差
                    lower_tol = self.get_cell_value(row, col_indices['lower_tol'])    # 第8列：下公差
                    measured = self.get_cell_value(row, col_indices['measured'])        # 第6列：实测值

                    # 创建键（尺寸+轴）
                    size_str = str(size).strip()
                    axis_str = str(axis).strip()
                    key = (size_str, axis_str)

                    # 如果是第一次遇到这个键，初始化数据
                    if key not in all_data:
                        all_data[key] = {
                            'nominal': nominal,
                            'upper_tol': upper_tol,
                            'lower_tol': lower_tol,
                            'measurements': {}
                        }

                    # 初始化模序号的测量值字典
                    if mold_num not in all_data[key]['measurements']:
                        all_data[key]['measurements'][mold_num] = {}

                    # 为当前尺寸+轴组合初始化穴数索引（如果尚未初始化）
                    if key not in size_cavity_indices:
                        size_cavity_indices[key] = start_cavity - 1  # 每个尺寸+轴组合的穴数索引都从start_cavity-1开始

                    # 添加实测值（第6列的数据），按照当前尺寸+轴组合的穴数索引存储
                    if measured is not None and not pd.isna(measured):
                        cavity_num = size_cavity_indices[key] + 1
                        all_data[key]['measurements'][mold_num][cavity_num] = measured
                        size_cavity_indices[key] += 1

            # 检查产品名是否一致
            if len(set(product_names)) > 1:
                return False, "警告：所有文件的产品名不一致！", None

            # 确定产品名
            product_name = product_names[0] if product_names else "三次元汇总"

            # 生成输出文件
            output_file = self.generate_three_d_output(all_data, output_dir, create_subfolder, product_name, layout_mode)

            return True, f"汇总完成，共处理 {len(file_paths)} 个文件 -> {os.path.basename(output_file)}", output_file

        except Exception as e:
            return False, f"处理失败: {str(e)}", None

    def parse_filename(self, filename):
        """解析文件名，提取产品名、模数和穴数范围"""
        try:
            # 移除文件扩展名
            filename_without_ext = os.path.splitext(filename)[0]
            
            # 提取产品名、模数和穴数范围
            # 分割字符串，找到"第n模"的位置
            parts = filename_without_ext.split('-')
            mold_part_index = -1
            for i, part in enumerate(parts):
                if '第' in part and '模' in part:
                    mold_part_index = i
                    break
            
            if mold_part_index == -1:
                return None
            
            # 产品名：从开始到"第n模"前面的部分
            product_name = '-'.join(parts[:mold_part_index])
            
            # 模数："第n模"部分
            mold_part = parts[mold_part_index]
            # 提取模序号
            mold_num = int(''.join(filter(str.isdigit, mold_part)))
            
            # 穴数范围："第n模"后面的部分
            cavity_range_str = '-'.join(parts[mold_part_index+1:])
            
            # 提取穴数范围的开始和结束值
            try:
                if '~' in cavity_range_str:
                    # 范围格式："1穴~16穴"
                    start_cavity = int(cavity_range_str.split('~')[0].replace('穴', ''))
                    end_cavity = int(cavity_range_str.split('~')[1].replace('穴', ''))
                    cavity_range = (start_cavity, end_cavity)
                else:
                    # 单个穴数格式："1穴"
                    cavity_num = int(cavity_range_str.replace('穴', ''))
                    cavity_range = (cavity_num, cavity_num)  # 开始和结束都是同一个值
            except:
                cavity_range = None
            
            return product_name, mold_num, cavity_range
        except:
            return None

    def find_data_start_row(self, df):
        """查找数据开始行（跳过标题行）"""
        # 通常第一行是标题，第二行开始是数据
        # 但我们需要检查第一行是否真的是标题
        for i in range(min(10, len(df))):  # 增加搜索范围到10行
            row = df.iloc[i]
            # 检查是否包含"尺寸"、"轴"等关键词
            row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)])
            row_str_upper = row_str.upper()
            
            # 如果包含标题关键词，下一行是数据
            if any(keyword in row_str for keyword in ['尺寸', '轴', '标准值', '上公差', '下公差', '实测值', '特征', '描述']) or \
               any(keyword in row_str_upper for keyword in ['SIZE', 'AXIS', 'NOMINAL', 'TOL', 'MEAS', 'FEATURE', 'DESCRIPTION']):
                        # 检查下一行是否是数据行（包含数字或有效数据）
                        if i + 1 < len(df):
                            next_row = df.iloc[i + 1]
                            # 检查下一行是否包含有效数据（前几列有值）
                            if len(next_row) >= 4:
                                # 检查第1列和第4列是否有值（尺寸、轴）
                                if pd.notna(next_row.iloc[0]) and pd.notna(next_row.iloc[3]):
                                    return i + 1  # 下一行开始是数据
                            elif len(next_row) >= 1:
                                # 如果列数不够，至少检查第一列
                                if pd.notna(next_row.iloc[0]):
                                    return i + 1  # 下一行开始是数据
                        return i + 1  # 即使下一行可能为空，也返回下一行
        
        # 如果没找到标题行，尝试从第一行开始查找数据
        # 检查第一行是否是数据行（包含数字）
        if len(df) > 0:
            first_row = df.iloc[0]
            if len(first_row) >= 4:
                # 检查第1列和第4列是否有值（尺寸、轴）
                if pd.notna(first_row.iloc[0]) and pd.notna(first_row.iloc[3]):
                    return 0  # 第一行就是数据
        
        # 默认从第二行开始（索引1），如果只有一行则从第一行开始
        return 1 if len(df) > 1 else 0

    def detect_column_indices(self, df, header_row):
        """检测列索引，使用固定的列索引映射"""
        # 根据用户要求，使用固定的列索引：
        # 汇总文件第1列（尺寸）：源文件第1列（索引0）
        # 汇总文件第2列（描述）：源文件第2列（索引1）
        # 汇总文件第3列（特征）：源文件第3列（索引2）
        # 汇总文件第4列（轴）：源文件第4列（索引3）
        # 汇总文件第5列（标准值NOMINAL）：源文件第5列（索引4）
        # 汇总文件第6列（上公差+TOL）：源文件第7列（索引6）
        # 汇总文件第7列（下公差-TOL）：源文件第8列（索引7）
        # 汇总文件第8列开始（实测值MEAS）：源文件第6列（索引5）
        
        col_indices = {
            'size': 0,        # 第1列：尺寸
            'description': 1,  # 第2列：描述
            'feature': 2,     # 第3列：特征
            'axis': 3,        # 第4列：轴
            'nominal': 4,     # 第5列：标准值
            'upper_tol': 6,   # 第7列：上公差
            'lower_tol': 7,   # 第8列：下公差
            'measured': 5     # 第6列：实测值
        }
        
        return col_indices

    def get_cell_value(self, row, col_idx):
        """获取单元格值"""
        if col_idx >= len(row):
            return None
        value = row.iloc[col_idx]
        return value if pd.notna(value) else None

    def generate_three_d_output(self, all_data, output_dir, create_subfolder, product_name, layout_mode="horizontal"):
        """生成汇总后的Excel文件"""
        # 准备输出数据
        output_rows = []

        # 收集所有的模序号和穴序号
        all_mold_nums = set()
        all_cavity_nums = set()
        for key, data in all_data.items():
            for mold_num in data['measurements']:
                all_mold_nums.add(mold_num)
                for cavity_num in data['measurements'][mold_num]:
                    all_cavity_nums.add(cavity_num)
        
        # 排序模序号和穴序号
        sorted_mold_nums = sorted(all_mold_nums)
        sorted_cavity_nums = sorted(all_cavity_nums)

        # 数据行（按尺寸结合轴排序）
        def get_sort_key(key):
            """生成排序键，按照尺寸列结合轴列排序"""
            size, axis = key
            size_str = str(size).strip()
            axis_str = str(axis).strip()
            
            # 初始化排序键
            size_sort_key = None
            axis_sort_key = None
            
            try:
                # 处理尺寸排序
                # 检查是否以数字开头
                if size_str and size_str[0].isdigit():
                    # 数字开头的标签
                    if '*' in size_str:
                        # 处理带星号的标签，如 "15*1", "15*2"
                        num_parts = size_str.split('*')
                        main_num = int(num_parts[0])
                        sub_num = int(num_parts[1]) if len(num_parts) > 1 else 0
                        size_sort_key = (0, '', main_num, sub_num)
                    elif '-' in size_str:
                        # 处理带连字符的标签，如 "FAI10-1"
                        # 提取前缀和数字部分
                        prefix = ''.join(filter(str.isalpha, size_str.split('-')[0]))
                        num_part = size_str.split('-')[0].replace(prefix, '')
                        suffix_part = size_str.split('-')[1] if len(size_str.split('-')) > 1 else '0'
                        try:
                            main_num = int(num_part)
                            sub_num = int(suffix_part)
                        except ValueError:
                            main_num = 0
                            sub_num = 0
                        # 对于FAI开头的，放在字母开头的后面
                        if prefix.upper() == 'FAI':
                            size_sort_key = (2, prefix.upper(), main_num, sub_num)
                        else:
                            size_sort_key = (0, prefix, main_num, sub_num)
                    else:
                        # 处理纯数字标签，如 "1", "2", "11"
                        main_num = int(size_str)
                        size_sort_key = (0, '', main_num, 0)
                else:
                    # 字母开头的标签
                    # 提取字母部分和数字部分
                    alpha_part = ''
                    num_part = ''
                    star_part = ''
                    
                    # 找到第一个数字的位置
                    for i, char in enumerate(size_str):
                        if char.isdigit():
                            alpha_part = size_str[:i]
                            remaining = size_str[i:]
                            # 处理可能的星号
                            if '*' in remaining:
                                num_star_parts = remaining.split('*')
                                num_part = num_star_parts[0]
                                star_part = num_star_parts[1] if len(num_star_parts) > 1 else '0'
                            elif '-' in remaining:
                                # 处理带连字符的标签，如 "E10-1"
                                num_part = remaining.split('-')[0]
                                star_part = remaining.split('-')[1] if len(remaining.split('-')) > 1 else '0'
                            else:
                                num_part = remaining
                                star_part = '0'
                            break
                    
                    # 转换数字部分为整数
                    try:
                        num_val = int(num_part)
                        star_val = int(star_part)
                    except ValueError:
                        num_val = 0
                        star_val = 0
                    
                    # 字母开头的排序键：(1, 字母部分, 数字部分, 星号部分)
                    size_sort_key = (1, alpha_part.upper(), num_val, star_val)
            except (ValueError, IndexError):
                # 如果解析失败，放到最后
                size_sort_key = (3, '', 0, 0)
            
            # 处理轴排序：先字母后汉字
            try:
                # 检查轴是否以字母开头
                if axis_str and (axis_str[0].isalpha() or axis_str[0] in 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'):
                    # 字母开头的轴
                    axis_sort_key = (0, axis_str)
                else:
                    # 汉字或其他开头的轴
                    axis_sort_key = (1, axis_str)
            except (ValueError, IndexError):
                # 如果解析失败，放到最后
                axis_sort_key = (2, axis_str)
            
            # 组合排序键：(尺寸排序键, 轴排序键)
            return (size_sort_key, axis_sort_key)
        
        sorted_keys = sorted(all_data.keys(), key=get_sort_key)

        if layout_mode == "horizontal":
            # 横排模式：每3穴换行输出
            # 分组穴数
            group_size = 3
            
            # 计算分组数
            num_cavities = len(sorted_cavity_nums)
            num_groups = (num_cavities + group_size - 1) // group_size
            
            # 对穴数进行分组
            cavity_groups = []
            for i in range(num_groups):
                start_idx = i * group_size
                end_idx = min(start_idx + group_size, num_cavities)
                cavity_groups.append(sorted_cavity_nums[start_idx:end_idx])
            
            # 为每个分组生成输出
            for group_idx, cavity_group in enumerate(cavity_groups):
                # 标题行（去除"描述"列）
                header = ['尺寸', '轴', 'NOMINAL', '+TOL', '-TOL']
                
                # 添加实测值列标题
                for cavity_num in cavity_group:
                    for mold_num in sorted_mold_nums:
                        header.append(f'{cavity_num}-#{mold_num}')
                
                output_rows.append(header)
                
                # 为每个尺寸生成数据行
                for key in sorted_keys:
                    size, axis = key
                    data = all_data[key]

                    row = [
                        size,  # 第1列：尺寸（源文件第1列）
                        axis,  # 第2列：轴（源文件第4列）
                        data['nominal'] if pd.notna(data['nominal']) else '',  # 第3列：NOMINAL（源文件第5列）
                        data['upper_tol'] if pd.notna(data['upper_tol']) else '',  # 第4列：+TOL（源文件第7列）
                        data['lower_tol'] if pd.notna(data['lower_tol']) else ''  # 第5列：-TOL（源文件第8列）
                    ]

                    # 添加实测值
                    for cavity_num in cavity_group:
                        for mold_num in sorted_mold_nums:
                            # 查找对应模序号和穴序号的测量值
                            if mold_num in data['measurements'] and cavity_num in data['measurements'][mold_num]:
                                row.append(data['measurements'][mold_num][cavity_num])
                            else:
                                row.append('')

                    output_rows.append(row)
                
                # 在分组之间添加空行（最后一组除外）
                if group_idx < num_groups - 1:
                    output_rows.append([''] * len(header))
        else:
            # 竖排模式：按模次分组，第6列为模次，第7列开始为各穴实测值
            # 标题行（去除"描述"列）
            header = ['尺寸', '轴', 'NOMINAL', '+TOL', '-TOL', '模次']
            
            # 添加实测值列标题（穴数）
            for cavity_num in sorted_cavity_nums:
                header.append(f'{cavity_num}穴')

            output_rows.append(header)

            # 为每个尺寸生成数据行（按模次分组）
            for key in sorted_keys:
                size, axis = key
                data = all_data[key]

                # 为每个模次生成一行数据
                for mold_num in sorted_mold_nums:
                    row = [
                        size,  # 第1列：尺寸（源文件第1列）
                        axis,  # 第2列：轴（源文件第4列）
                        data['nominal'] if pd.notna(data['nominal']) else '',  # 第3列：NOMINAL（源文件第5列）
                        data['upper_tol'] if pd.notna(data['upper_tol']) else '',  # 第4列：+TOL（源文件第7列）
                        data['lower_tol'] if pd.notna(data['lower_tol']) else '',  # 第5列：-TOL（源文件第8列）
                        f'#{mold_num}'  # 第6列：模次
                    ]

                    # 添加各穴实测值
                    for cavity_num in sorted_cavity_nums:
                        # 查找对应模序号和穴序号的测量值
                        if mold_num in data['measurements'] and cavity_num in data['measurements'][mold_num]:
                            row.append(data['measurements'][mold_num][cavity_num])
                        else:
                            row.append('')

                    output_rows.append(row)

        # 创建DataFrame
        output_df = pd.DataFrame(output_rows)

        # 生成输出文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 确定输出路径
        if create_subfolder:
            subfolder_name = "三次元数据汇总"
            output_folder = os.path.join(output_dir, subfolder_name)
            os.makedirs(output_folder, exist_ok=True)
        else:
            output_folder = output_dir
        
        # 使用"产品名+横排/竖排+时间"格式命名输出文件
        layout_prefix = "横排" if layout_mode == "horizontal" else "竖排"
        output_filename = f"{product_name}_{layout_prefix}_{timestamp}.xlsx"
        output_file = os.path.join(output_folder, output_filename)

        # 保存为Excel文件
        output_df.to_excel(output_file, index=False, header=False, engine='openpyxl')

        # 加载Excel文件并设置标题行和尺寸列加粗，根据布局模式决定是否冻结
        try:
            wb = load_workbook(output_file)
            ws = wb.active

            # 根据布局模式设置不同的冻结列
            if layout_mode == "horizontal":
                # 横排模式：冻结F列（使用F1单元格作为冻结点）
                ws.freeze_panes = 'F1'
            else:
                # 竖排模式：冻结G列（使用G1单元格作为冻结点）
                ws.freeze_panes = 'G1'
            
            # 设置标题行文字加粗和尺寸列加粗
            from openpyxl.styles import Font
            bold_font = Font(bold=True)
            
            # 加粗标题行
            if layout_mode == "horizontal":
                # 横排模式：每个分组的第一行都是标题行
                for row in ws.iter_rows():
                    if row[0].value == "尺寸":
                        for cell in row:
                            cell.font = bold_font
            else:
                # 竖排模式：只有第一行是标题行
                for cell in ws[1]:
                    cell.font = bold_font
            
            # 加粗尺寸列（第一列）
            for row in ws.iter_rows():
                if row[0].value and row[0].value != "尺寸":  # 排除标题行的"尺寸"单元格
                    row[0].font = bold_font
            
            # 保存修改
            wb.save(output_file)
        except Exception as e:
            print(f"设置冻结窗格或加粗失败: {e}")

        return output_file
