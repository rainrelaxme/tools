# -*- coding:utf-8 -*-
# authored by rainrelaxme
# 说明：文件工具的中间过程及最终函数实现

import os
import xlsxwriter  # 只能写入xlsx，如果要写入xls要用xlwt
from openpyxl import Workbook

from office_tools import file_edit
from office_tools import excel_edit


def get_filelist(folder_path, output_path=os.path.join(os.path.expanduser('~'), 'Downloads'), mode=0):
    """获取文件夹下的文件清单，输出到excel，序号、文件名、大小、后缀名,
        output_path:输出文件夹默认，默认“下载”。
        mode: 0:不返回下层文件夹，1：返回下层文件夹
    """
    if mode == 0:
        # 创建一个新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "File Names"

        # 写入表头
        ws['A1'] = "File Name"

        # 遍历文件夹中的所有文件
        for idx, filename in enumerate(os.listdir(folder_path), start=2):
            # 只处理文件，忽略子文件夹
            if os.path.isfile(os.path.join(folder_path, filename)):
                ws[f'A{idx}'] = filename

        # 保存Excel文件
        wb.save(output_path)
        print(f"File names have been saved to {output_path}")


def rename_add_space(path):
    """文件名称增加空格"""
    file_list = os.listdir(path)
    for file in file_list:
        absolute_file = path + "/" + file
        if os.path.isfile(absolute_file):  # 不处理文件夹
            if file.endswith(".py") or file.endswith(".exe") or file.endswith(".txt") or file.endswith(".ipynb"):
                continue
            else:
                old_name = absolute_file
                forward_name = file_edit.split_name(file)[0]
                forward_name = file_edit.add_space(forward_name)
                back_name = file_edit.split_name(file)[1]
                new_name = path + "/" + forward_name + back_name
                os.rename(old_name, new_name)
        else:
            # 递归文件夹修改文件名
            rename_add_space(absolute_file)


def rename_sub_space(path):
    """文件名称去除空格"""
    file_list = os.listdir(path)
    for file in file_list:
        absolute_file = path + "/" + file
        if os.path.isfile(absolute_file):  # 不处理文件夹
            if file.endswith(".py") or file.endswith(".exe") or file.endswith(".txt") or file.endswith(".ipynb"):
                continue
            else:
                old_name = absolute_file
                forward_name = file_edit.split_name(file)[0]
                forward_name = file_edit.sub_space1(forward_name)
                back_name = file_edit.split_name(file)[1]
                new_name = path + "/" + forward_name + back_name
                os.rename(old_name, new_name)
        else:
            # 递归文件夹修改文件名
            rename_sub_space(absolute_file)


def rename_add_x(path, x):
    """给当前路径下的所有文件名后边都加“x”"""
    file_list = os.listdir(path)
    for file in file_list:
        absolute_file = path + "/" + file
        if os.path.isfile(absolute_file):  # 不处理文件夹
            if file.endswith(".py") or file.endswith(".exe") or file.endswith(".txt") or file.endswith(".ipynb"):
                continue
            else:
                old_name = absolute_file
                forward_name = file_edit.split_name(file)[0]
                forward_name = file_edit.add_x(forward_name, x)
                back_name = file_edit.split_name(file)[1]
                new_name = path + "/" + forward_name + back_name
                os.rename(old_name, new_name)
        else:
            # 递归文件夹修改文件名
            rename_add_x(absolute_file, x)


def rename_sub_x(path, x):
    """给当前路径下的所有文件名后边都去除“x”"""
    file_list = os.listdir(path)
    for file in file_list:
        absolute_file = path + "/" + file
        if os.path.isfile(absolute_file):  # 不处理文件夹
            if file.endswith(".py") or file.endswith(".exe") or file.endswith(".txt") or file.endswith(".ipynb"):
                continue
            else:
                old_name = absolute_file
                forward_name = file_edit.split_name(file)[0]
                forward_name = file_edit.sub_x(forward_name, x)
                back_name = file_edit.split_name(file)[1]
                new_name = path + "/" + forward_name + back_name
                os.rename(old_name, new_name)
        else:
            # 递归文件夹修改文件名
            rename_sub_x(absolute_file, x)


def format_name1(path, girl):
    """rename1(name, girl)"""
    pass


def format_name2(path, forward_name, start_num: int):
    """rename2(name, forward_name, start_num: int)"""
    pass


def excel_merge(files: tuple):
    """利用Python将多个excel文件合并为一个文件"""
    # 存储所有读取的结果
    value = []
    for file in files:
        f_open = excel_edit.open_xls(file)
        sh_num = excel_edit.get_sheet_num(f_open)
        for sheet in range(sh_num):
            print("正在读取文件：" + str(file) + "的第" + str(sheet + 1) + "个sheet表的内容...")
            sheet_value = excel_edit.get_file(file, sheet)
            if sheet_value:
                if value:
                    # 第一个表格直接读取，后续表格判断表头是否相同，不同不允许合并，相同则把表头去除
                    if sheet_value[0] != value[0]:
                        return "Error：表头不同，无法合并"
                    else:
                        for row in sheet_value[1:len(sheet_value) - 1]:
                            value.append(row)
                else:
                    for row in sheet_value:
                        value.append(row)
    # 定义最终合并后生成的新文件
    filetypes = [("Microsoft Excel文件", "*.xlsx"),
                 ("Microsoft Excel文件", "*.xls"),
                 ("CSV文件", "*.csv")]
    result_file = excel_edit.save_file(filetypes=filetypes, defaultextension=filetypes[0][1])
    work_file = xlsxwriter.Workbook(result_file.name)  # 创建工作表
    work_sheet = work_file.add_worksheet()  # 默认创建sheet1
    for row in range(len(value)):
        for col in range(len(value[row])):
            cell = value[row][col]
            work_sheet.write(row, col, cell)  # 向行列对应的单元格写入内容
    work_file.close()
    print("ok")


def is_exist(filename, path, mode=0):
    """
    判断文件是否已经存在，mode=0精确匹配，mode=1模糊匹配
    如果已经存在，则返回1；不存在，则返回0.
    """
    file_list = os.listdir(path)
    results = []
    if mode == 0:
        if filename in file_list:
            results.append(filename)
        else:
            pass
    elif mode == 1:
        for file in file_list:
            result = re.search(filename, file)
            if result:
                results.append(result.string)
            else:
                continue
    else:
        print('Error mode!')
    if results:
        print(results)
        return 0, '\n', results
    else:
        return 1


# all_file = ('E:/project/pythonProject/Little_tools/src/1.xlsx', 'E:/project/pythonProject/Little_tools/src/2.xlsx')
# all_file = ('E:/project/pythonProject/Little_tools/src/3.xlsx')
# excel_merge(all_file)

def test():
    get_filelist(r'C:\Users\shawn\Desktop\languages',r"D:\新建文件夹")


test()
