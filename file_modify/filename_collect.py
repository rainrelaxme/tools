#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
@Project: distance.py
@Version:
@Author: RainRelaxMe
@Date: 2025/3/10 21:36
"""


import os
from openpyxl import Workbook

def extract_filenames_to_excel(folder_path, output_excel_path):
    # 创建一个新的工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "File Names"

    # 写入表头
    ws['A1'] = "File Name"

    # 遍历文件夹中的所有文件
    for idx, filename in enumerate(os.listdir(folder_path), start=2):
        # 只处理文件，忽略子文件夹
        if os.path.isfile(os.path.join(folder_path, filename)):
            ws[f'A{idx}'] = filename

    # 保存Excel文件
    wb.save(output_excel_path)
    print(f"File names have been saved to {output_excel_path}")

# 示例用法
folder_path = "J:\待定\木田彩水"  # 替换为你的文件夹路径
output_excel_path = "output_file_names.xlsx"  # 替换为你想保存的Excel文件路径
extract_filenames_to_excel(folder_path, output_excel_path)