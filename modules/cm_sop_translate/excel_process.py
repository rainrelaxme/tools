#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
@Project: translate_tools.py
@File   : excel_process.py
@Version:
@Author : shawn
@Date   : 2025/12/1 13:00
@Info   : 实现excel文档的翻译，并用翻译后的文本替换原文本后，且保持原文档格式。
"""
import copy
import datetime
import io
import logging
import os
import shutil
import time
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pythoncom
from PIL import Image, ImageGrab
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from win32com import client as wc

from modules.cm_sop_translate.config.config import config
from modules.cm_sop_translate.translator import Translator
from modules.common.log import setup_logger

logger = setup_logger(log_dir=config.LOG_PATH, name='logs', level=logging.INFO)
TEMP_PATH = config.TEMP_PATH


def get_content(file_path: str) -> List[Dict[str, List[List[Optional[str]]]]]:
    """
    读取excel文件的内容，返回每个sheet的二维数组，保留原始顺序
    """
    workbook = load_workbook(file_path, data_only=False)
    data: List[Dict[str, List[List[Optional[str]]]]] = []

    # 设定最大的取值范围，防止空的单元格，但是被误编辑的被统计在内
    max_row = ''
    max_column = ''

    for sheet in workbook.worksheets:
        sheet_matrix: List[List[Optional[str]]] = []
        cells_matrix = []
        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                # row_values.append(cell.value)
                # 获取单元格格式-文字
                cell_info = {
                    "type": '',
                    "flag": '',
                    "is_merged": False,
                    "is_merge_start": False,
                    "coordinate": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "value": cell.value if cell.value else "",
                    "format": get_cell_format(cell)
                }
                # print(cell_info)
                cells_matrix.append(cell_info)
            # sheet_matrix.append(row_values)
        # 合并单元格信息
        merge_info = get_merge_info(sheet)

        sheet_data = {
            "sheet": sheet.title,
            "sheet_color": sheet.sheet_properties.tabColor.rgb if sheet.sheet_properties.tabColor else None,
            # "rows": sheet_matrix,
            # "cells": cells_matrix,
            "cells": update_merge_info(cells_matrix, merge_info),
            "merge_info": merge_info,
            "shape_info": extract_shape_from_excel(file_path, sheet.title, TEMP_PATH)
        }
        data.append(sheet_data)

    return data


def get_row_format(file_path, row_number, sheet_name=None):
    """
    获取整行的格式信息
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    row_format = {
        "row_height": ws.row_dimensions[row_number].height if row_number in ws.row_dimensions else None,
        "cells": {}
    }

    # 获取每列的格式
    max_column = ws.max_column
    for col in range(1, max_column + 1):
        cell = ws.cell(row=row_number, column=col)
        row_format["cells"][cell.column_letter] = get_cell_format(cell)

    return row_format


def get_column_format(file_path, column_letter, sheet_name=None):
    """
    获取整列的格式信息
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    col_idx = ws[column_letter + "1"].column

    column_format = {
        "column_letter": column_letter,
        "column_width": ws.column_dimensions[column_letter].width if column_letter in ws.column_dimensions else None,
        "cells": {}
    }

    # 获取每行的格式
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        column_format["cells"][row] = get_cell_format(cell)

    return column_format


def get_cell_format(cell):
    """
    获取单元格的格式内容
    """
    cell_format = {
        "value": cell.value,
        "data_type": type(cell.value).__name__,
        "number_format": cell.number_format,
        "font": {},
        "fill": {},
        "border": {},
        "alignment": {},
        "protection": {}
    }

    # 1. 获取字体格式
    if cell.font:
        font: Font = cell.font
        # 如果有颜色的属性，则可以获取到
        color = None
        color_type = None
        if font.color:
            color = font.color.rgb if font.color.type == "rgb" else font.color.theme
            color_type = font.color.type

        cell_format["font"] = {
            "name": font.name,
            "size": font.sz,
            "bold": font.b,
            "italic": font.i,
            "underline": font.u,
            "strike": font.strike,
            "color": color,
            "color_type": color_type
        }

    # 2. 获取填充格式（背景色）
    if cell.fill:
        fill: PatternFill = cell.fill
        cell_format["fill"] = {
            "fill_type": fill.fill_type,
            "start_color": fill.start_color.rgb if fill.start_color.type == "rgb" else fill.start_color.theme,
            "end_color": fill.end_color.rgb if fill.end_color.type == "rgb" else fill.end_color.theme,
            "fgColor": fill.fgColor.rgb if hasattr(fill.fgColor, 'rgb') else str(fill.fgColor),
            "bgColor": fill.bgColor.rgb if hasattr(fill.bgColor, 'rgb') else str(fill.bgColor)
        }

    # 3. 获取边框格式
    if cell.border:
        border: Border = cell.border
        cell_format["border"] = {
            "left": get_side_info(border.left),
            "right": get_side_info(border.right),
            "top": get_side_info(border.top),
            "bottom": get_side_info(border.bottom),
            "diagonal": get_side_info(border.diagonal),
            "diagonal_direction": border.diagonal_direction,
            "outline": border.outline  # 外边框
        }

    # 4. 获取对齐格式
    if cell.alignment:
        alignment: Alignment = cell.alignment
        cell_format["alignment"] = {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "text_rotation": alignment.textRotation,
            "wrap_text": alignment.wrapText,
            "shrink_to_fit": alignment.shrinkToFit,
            "indent": alignment.indent
        }

    # 5. 获取保护格式
    if cell.protection:
        protection: Protection = cell.protection
        cell_format["protection"] = {
            "locked": protection.locked,
            "hidden": protection.hidden
        }

    return cell_format


def get_side_info(side: Side):
    """获取边框边的信息"""
    if side and side.style:
        return {
            "style": side.style,
            "color": side.color.rgb if side.color and side.color.type == "rgb" else side.color.theme if side.color else None
        }
    return None


def get_merge_info(sheet):
    """
    获取合并单元格信息
    """
    merge_lists = sheet.merged_cells
    # print("merge_list*******", merge_lists)

    merge_infos = []
    for merge_list in merge_lists:
        # 一种方法
        merge_coords = merge_list.coord

        # 另一种方法可以把所有合并的单元格返回
        # 获取合并单元格的起始行列，和终止行列
        row_min, row_max, col_min, col_max = merge_list.min_row, merge_list.max_row, merge_list.min_col, merge_list.max_col
        merge_start_end = [(row_min, col_min), (row_max, col_max)]
        # 定义合并单元格的类型：合并行列/合并行/合并列
        merge_type = ''
        row_col = []
        if row_min != row_max and col_min != col_max:  # 合并行列
            row_col = [(x, y) for x in range(row_min, row_max + 1) for y in range(col_min, col_max + 1)]
            merge_type = 'row_col_merge'
        elif row_min == row_max and col_min != col_max:  # 合并行
            row_col = [(row_min, y) for y in range(col_min, col_max + 1)]
            merge_type = 'row_merge'
        elif row_min != row_max and col_max == col_min:  # 合并行
            row_col = [(x, col_min) for x in range(row_min, row_max + 1)]
            merge_type = 'col_merge'

        merge_info = {
            "merge_coords": merge_coords,
            "merge_start_end": merge_start_end,  # 合并单元格的起止位置
            "merge_cells": row_col,  # 合并的单元格序号
            "merge_type": merge_type,  # 合并单元格的类型
            "merge_data": sheet.cell(row=merge_list.min_row, column=merge_list.min_col).value,  # 合并单元格的内容
        }
        merge_infos.append(merge_info)

    return merge_infos


def update_merge_info(data, merge_info):
    """
    标记单元格是否是合并单元格，合并单元格除左上角外禁止写入
    """
    new_data = copy.deepcopy(data)

    # 记录所有合并单元格的左上角的单元格list
    merge_start_list = []
    # 记录所有合并单元格的单元格list
    merge_cell_list = []
    for merge_block in merge_info:
        start_cell = merge_block["merge_coords"].split(":", 1)[0]
        merge_start_list.append(start_cell)
        merge_cell_list.extend(merge_block["merge_cells"])

    # 更改is_merged和is_merge_start的值
    for cell in new_data:
        # cell_pos = (cell['row'], cell['col'])
        if cell['coordinate'] in merge_start_list:
            cell['is_merge_start'] = True
            cell['is_merged'] = True
        elif (cell['row'], cell['col']) in merge_cell_list:
            cell['is_merge_start'] = False
            cell['is_merged'] = True

    return new_data


def extract_images_from_excel(file_path, output_dir='extracted_images'):
    """
    从Excel文件中提取图片
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook = load_workbook(file_path)
    image_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        image_data[sheet_name] = []

        # 获取工作表的所有图片
        for img in sheet._images:
            # 获取图片数据
            img_data = img._data()

            # 创建图片对象
            image = Image.open(io.BytesIO(img_data))

            # 生成唯一文件名
            import uuid
            img_filename = f"{sheet_name}_{uuid.uuid4().hex[:8]}.png"
            img_path = os.path.join(output_dir, img_filename)

            # 保存图片
            image.save(img_path)

            # 记录图片位置信息
            image_info = {
                'path': img_path,
                'anchor': img.anchor,
                'filename': img_filename
            }
            image_data[sheet_name].append(image_info)

    return image_data


def extract_shape_from_excel(excel_path, sheet_name, output_folder):
    """
    将Excel工作表中重叠的图片和形状组合并导出为图像

    参数:
    excel_path: Excel文件路径
    sheet_name: 工作表名称
    output_image_path: 输出图像文件路径
    overlap_threshold: 重叠阈值(0-1)，默认30%重叠视为需要组合
    """
    excelApp = None
    workbook_win32 = None
    shapes_info = []
    # 确保输出文件夹存在
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    try:
        # 初始化COM (单线程公寓模式)
        # pythoncom.CoInitialize()

        # 启动Excel应用程序
        excelApp = wc.Dispatch("Excel.Application")
        excelApp.Visible = False  # 不可见模式运行
        excelApp.DisplayAlerts = False  # 不显示警告
        excelApp.ScreenUpdating = False  # 禁用屏幕更新以提高性能

        logger.info(f"正在打开Excel文件: {excel_path}, 获取{sheet_name}的内容。")

        # 打开工作簿
        workbook_win32 = excelApp.Workbooks.Open(os.path.abspath(excel_path))
        worksheet = workbook_win32.Worksheets(sheet_name)

        # 激活工作表
        worksheet.Activate()

        # 获取所有形状
        shapes = worksheet.Shapes

        # 如果没有形状，直接返回
        if shapes.Count == 0:
            logger.warning("工作表中没有找到形状或图片")
            return

        logger.info(f"找到 {shapes.Count} 个形状/图片")

        # 收集所有形状的信息
        for i in range(1, shapes.Count + 1):
            try:
                shape = shapes.Item(i)
                text = ''
                try:
                    # if hasattr(shape, 'TextEffect'):
                    if shape.Type == 17:
                        text = shape.TextEffect.Text
                except Exception as e:
                    logger.error(str(e))

                shape_info = {
                    'id': shape.ID,
                    'name': shape.Name,
                    'type_id': shape.Type,
                    "type": get_shape_type(shape),
                    'left': shape.Left,
                    'top': shape.Top,
                    'width': shape.Width,
                    'height': shape.Height,
                    # 'shape': shape,
                    'text': text
                }

                # # 复制形状
                # shape.Copy()
                # # 等待剪贴板数据到来
                # time.sleep(0.5)
                # # 从剪贴板获取图像
                # image = ImageGrab.grabclipboard()
                #
                # if image is not None:
                #     # CT = datetime.datetime.now().strftime('%y%m%d%H%M%S')
                #     output_pic = os.path.join(output_folder, f"{sheet_name}_{shape.Name}.png")
                #     image.save(output_pic, 'PNG')
                #     shape_info["path"] = output_pic
                #     print(f"图片保存成功{output_pic}")
                # else:
                #     print("未能从剪贴板获取图像")

                shapes_info.append(shape_info)
            except Exception as e:
                logger.warning(f"无法获取形状 {i} 的信息: {e}")
                continue

        logger.info(f"shape_info: {shapes_info}")

    except Exception as e:
        logger.error(f"处理过程中发生错误: {e}")
        # 获取更详细的错误信息
        try:
            import win32api
            win32api.FormatMessage(e.excepinfo[5])
        except:
            pass
        raise e

    finally:
        # 确保清理资源
        try:
            if workbook_win32:
                workbook_win32.Close(SaveChanges=False)
                logger.info("工作簿已关闭")
        except Exception as e:
            logger.error(f"关闭工作簿时出错: {e}")

        try:
            if excelApp:
                excelApp.Quit()
                logger.info("Excel应用程序已退出")
        except Exception as e:
            logger.error(f"退出Excel时出错: {e}")

        # 释放COM资源
        # pythoncom.CoUninitialize()

        return shapes_info


def get_shape_type(shape):
    """
    判断形状类型
    """
    shape_type = shape.Type
    if shape.Type == 1 and shape.Name.startswith("Rectangle"):
        shape_type = "Rectangle"
    if shape.Type == 13 and shape.Name.startswith("Picture"):
        shape_type = "Picture"
    # Group需要组合
    if shape.Type == 6 and shape.Name.startswith("Group"):
        shape_type = "Group"
    # TextBox需要翻译
    if shape.Type == 17 and shape.Name.startswith("TextBox"):
        shape_type = "TextBox"
    return shape_type


def paste_shape_to_excel(input_file, output_file, sheet_name):
    """
    绘制形状：textbox
    采用copy的方法
    """
    excelApp = wc.Dispatch("Excel.Application")
    excelApp.Visible = False

    workbook = excelApp.Workbooks.Open(input_file)
    sheet = workbook.Worksheets("Sheet1")

    new_workbook = excelApp.Workbooks.Open(output_file)
    new_sheet = new_workbook.Worksheets(sheet_name)

    for shape in sheet.Shapes:
        shape.Copy()
        new_sheet.Paste()

    new_workbook.SaveAs(output_file)
    print(f"{output_file}已保存成功！")
    excelApp.Quit()


def add_translation(data, translator, langs: list, method: str):
    """
    添加翻译段落
    method-replace: 替换原文字，只保留1种语言
    method-add: 增加新的内容，保留3种语言
    """
    tran_data = []
    if method == 'replace':  # 替换成一种语言
        for lang in langs:
            lang_data = copy.deepcopy(data)  # 使用copy()会影响原始数据，deepcopy不会影响
            for sheet_data in lang_data:
                for cell_data in sheet_data['cells']:
                    original_text = str(cell_data['value'])
                    if original_text.strip():
                        translated_text = translator.translate(original_text, lang, display=True)
                        cell_data['value'] = translated_text
            tran_data.append({
                "language": lang,
                "data": lang_data
            })

    elif method == 'replace_multi':  # 替换成多种语言
        # lang_data = copy.deepcopy(data)
        for sheet_data in data:
            # 翻译单元格文字
            for cell_data in sheet_data['cells']:
                original_text = str(cell_data['value'])
                if original_text.strip():
                    all_translated_text = original_text
                    for lang in langs:
                        translated_text = translator.translate(original_text, lang, display=True)
                        all_translated_text = all_translated_text + '\n' + translated_text
                    cell_data['value'] = all_translated_text
            # 翻译shape文本框文字
            for shape_data in sheet_data['shape_info']:
                original_text = str(shape_data['text'])
                if original_text.strip():
                    all_translated_text = original_text
                    for lang in langs:
                        translated_text = translator.translate(original_text, lang, display=True)
                        all_translated_text = all_translated_text + '\n' + translated_text
                    shape_data['text'] = all_translated_text

        tran_data.append({
            "language": '&'.join(langs),
            "data": data
        })

    elif method == 'add':
        pass

    return tran_data


def apply_cell_format(cell, data):
    """
    应用单元格的样式
    """
    #  字体
    cell.font = Font(
        name=data['format']['font']['name'],
        sz=data['format']['font']['size'],
        b=data['format']['font']['bold'],
        i=data['format']['font']['italic'],
        u=data['format']['font']['underline'],
        strike=data['format']['font']['strike'],
        # color=data['format']['font']['color']
    )

    #  填充
    cell.fill = PatternFill(
        fill_type=data['format']['fill']['fill_type'],
        start_color=data['format']['fill']['start_color'],
        end_color=data['format']['fill']['end_color'],
        fgColor=data['format']['fill']['fgColor'],
        bgColor=data['format']['fill']['bgColor']
    )

    #  边框
    # 暂时未包含颜色，如color=data['format']['border']['left']['color']
    cell.border = Border(
        left=Side(
            style=data['format']['border']['left'].get('style') if data['format']['border'].get('left') else None),
        right=Side(
            style=data['format']['border']['right'].get('style') if data['format']['border'].get('right') else None),
        top=Side(style=data['format']['border']['top'].get('style') if data['format']['border'].get('top') else None),
        bottom=Side(
            style=data['format']['border']['bottom'].get('style') if data['format']['border'].get('bottom') else None),
        diagonal=Side(
            style=data['format']['border'].get('diagonal') if data['format']['border'].get('diagonal') else None),
        diagonal_direction=Side(
            style=data['format']['border'].get('diagonal_direction') if data['format']['border'].get(
                'diagonal_direction') else None),
        # outline=Side(style=data['format']['border']['outline'])
    )

    #  对齐
    cell.alignment = Alignment(
        horizontal=data['format']['alignment']['horizontal'],
        vertical=data['format']['alignment']['vertical'],
        text_rotation=data['format']['alignment']['text_rotation'],
        wrap_text=data['format']['alignment']['wrap_text'],
        shrink_to_fit=data['format']['alignment']['shrink_to_fit'],
        indent=data['format']['alignment']['indent']
    )

    #  数字格式
    cell.number_format = data['format']['number_format']


# def remove_shape_from_excel(input_file, *shape, sheet_name: List = None, app=None):
#     """
#     去除excel中指定的shape类型
#     """
#     # 打开excel
#     # excelApp = wc.Dispatch("Excel.Application")
#     # excelApp.Visible = False  # 不可见模式运行
#     # excelApp.DisplayAlerts = False  # 不显示警告
#     # excelApp.ScreenUpdating = False  # 禁用屏幕更新以提高性能
#     excelApp = app
#     workbook = excelApp.Workbooks.Open(input_file)
#
#     # 获取所有待处理的sheet
#     if sheet_name:
#         all_sheets = sheet_name
#     else:
#         all_sheets = [sheet.Name for sheet in workbook.Sheets]
#
#     for sht_name in all_sheets:
#         sheet = workbook.Worksheets(sht_name)
#         if shape:
#             # 移除指定的形状
#             for shape_type in shape:
#                 for shp in sheet.Shapes:
#                     if shp.Type == int(shape_type):
#                         shp.Delete()
#         else:
#             # 移除所有形状
#             for shp in sheet.Shapes:
#                 shp.Delete()
#
#     # 存储文件
#     output_file = input_file.replace('.xlsx', '_remove_shape.xlsx')
#     workbook.SaveAs(output_file)
#     print(f"{output_file}已保存成功！")
#     # excelApp.Quit()
#
#     return output_file


# def insert_shape_to_excel(input_file, output_file, data, method='draw', app=None):
#     """
#     将shape插入到excel中
#     """
#     # 打开excel
#     # excelApp = wc.Dispatch("Excel.Application")
#     # excelApp.Visible = False  # 不可见模式运行
#     # excelApp.DisplayAlerts = False  # 不显示警告
#     # excelApp.ScreenUpdating = False  # 禁用屏幕更新以提高性能
#     excelApp = app
#     workbook = excelApp.Workbooks.Open(input_file)
#     target_workbook = excelApp.Workbooks.Open(output_file)
#
#     # 从已经保存png文件添加到excel中
#     if method == 'draw':
#         for sheet_data in data:
#             sheet = workbook.Worksheets(sheet_data['sheet'])
#             for shape_data in sheet_data['shape_info']['shape']:
#                 img_path = f"{sheet_data['shape_info']['shape_path']}/{sheet_data['sheet']}_{shape_data['name']}.png"
#                 img = sheet.Pictures().Insert(img_path)
#                 img.Left = shape_data['left']
#                 img.Top = shape_data['top']
#                 img.Width = shape_data['width']
#                 img.Height = shape_data['height']
#
#     # 从原excel中复制shape到新文件中
#     elif method == 'paste':
#         for sheet_data in data:
#             sheet = workbook.Worksheets(sheet_data['sheet'])
#             target_sheet = target_workbook.Worksheets(sheet_data['sheet'])
#             for shape_data in sheet_data['shape_info']:
#                 shape = sheet.Shapes(shape_data['name'])
#                 # 替换TextBox文本内容
#                 if shape.Type == 17:
#                     shape.TextEffect.Text = shape_data['text']
#                 shape.Copy()
#                 target_sheet.Paste()
#
#     # 存储文件
#     # new_output = output_file.replace('.xlsx', '_2.xlsx')
#     # target_workbook.SaveAs(new_output)
#     # print(f"{new_output}已保存成功！")
#     # excelApp.Quit()
#
#     return target_workbook


def create_new_excel(output_path: str, data, method, input_file=None):
    """
    根据内容创建新的excel文件
    方法-simple：直接替换value
    方法-complex：获取数据+格式等，再生成新的文件
    """
    # 建立win32应用
    excelApp = wc.Dispatch("Excel.Application")
    excelApp.Visible = False  # 不可见模式运行
    excelApp.DisplayAlerts = False  # 不显示警告
    excelApp.ScreenUpdating = False  # 禁用屏幕更新以提高性能

    # 方法2： 找到合并单元格的位置信息，只更新允许写入的内容
    if method == 'simple':
        for lang_data in data:
            lang = lang_data['language']

            # 替换单元格内容并保存新文件
            excel = openpyxl.load_workbook(input_file)
            for sheet_data in lang_data['data']:
                sheet = excel[sheet_data['sheet']]
                # 更换单元格内容
                for cell in sheet_data['cells']:
                    if cell['is_merged'] and not cell['is_merge_start']:
                        continue
                    elif cell['value'] == '':
                        continue
                    else:
                        sheet[cell['coordinate']] = cell['value']

            # 输出文件
            CT = datetime.datetime.now().strftime('%y%m%d%H%M%S')
            file_base_name = os.path.basename(input_file)
            output_file = f"{output_path}/{file_base_name.replace(".xlsx", f"_tran_{lang}_{CT}.xlsx")}"
            excel.save(output_file)

            # 替换形状
            replace_shape_on_excel(output_file, input_file, lang_data['data'], method='paste', app=excelApp)
            # 再次退出excel
            excelApp.Quit()

            return output_file

    elif method == 'copy':
        # 方法3： 整个sheet复制？不如直接复制整个excel
        # 有问题，替换单元格内容是用的openpyxl，一旦用她load_workbook，还是会丢失形状。不过理论上win32也可以替换单元格内容。。。
        pass

    elif method == 'simple2':  # 未使用此方法
        for lang_data in data:
            lang = lang_data['language']
            # 读取原文件
            excel = openpyxl.load_workbook(input_file)

            # 方法1：合并单元格，先取消再合并，以便写入内容 - 会丢失合并单元格中的图片信息，主要是形状信息
            for sheet_data in lang_data['data']:
                sheet = excel[sheet_data['sheet']]
                # 获取合并单元格信息
                merged_ranges = list(sheet.merged_cells.ranges)
                # 取消合并
                for merged_range in merged_ranges:
                    sheet.unmerge_cells(str(merged_range))
                # 更换单元格内容
                for cell in sheet_data['cells']:
                    sheet[cell['coordinate']] = cell['value']
                # 合并单元格
                for merged_range in merged_ranges:
                    sheet.merge_cells(str(merged_range))

            # 存储文件
            # new_output_file = output_file.replace(".xlsx", f"_{lang}.xlsx")
            # time.sleep(1)
            # excel.save(new_output_file)
            # print(f"输出文件：{new_output_file}")

    elif method == 'complex':
        excel = openpyxl.Workbook()
        # 清除已经存在的sheet
        default_sheet = excel.active
        excel.remove(default_sheet)

        # 创建新sheet并填充内容
        for sheet_data in data:
            sheet = excel.create_sheet(sheet_data['sheet'])  # sheet名称
            sheet.sheet_properties.tabColor = sheet_data['sheet_color']  # sheet颜色
            # 填充内容
            for cell in sheet_data['cells']:
                sheet[cell['coordinate']] = cell['value']
                apply_cell_format(sheet[cell['coordinate']], cell)

            # 合并单元格
            for merge_info in sheet_data['merge_info']:
                sheet.merge_cells(str(merge_info))

        # 存储文件
        # excel.save(output_file)
        # print(f"输出文件：{output_file}")


def replace_shape_on_excel(output_file, src_file, data, method='paste', *shape, sheets: List = None, app=None):
    """处理形状图片"""
    excelApp = app
    output_workbook = excelApp.Workbooks.Open(output_file)
    src_workbook = excelApp.Workbooks.Open(src_file)
    try:
        # 步骤1. 去除excel中的shape
        # 获取所有待处理的sheet
        if sheets:
            sheets = sheets
        else:
            sheets = [sheet.Name for sheet in output_workbook.Sheets]

        for sht_name in sheets:
            sheet = output_workbook.Worksheets(sht_name)
            if shape:
                # 移除指定的形状
                for shape_type in shape:
                    for shp in sheet.Shapes:
                        if shp.Type == int(shape_type):
                            shp.Delete()
            else:
                # 移除所有形状
                for shp in sheet.Shapes:
                    shp.Delete()

        # 步骤2. 添加形状到excel中
        # 方法1：从原excel中复制shape到新文件中
        if method == 'paste':
            for sheet_data in data:
                sheet = src_workbook.Worksheets(sheet_data['sheet'])
                target_sheet = output_workbook.Worksheets(sheet_data['sheet'])
                for shape_data in sheet_data['shape_info']:
                    shape = sheet.Shapes(shape_data['name'])
                    # 替换TextBox文本内容
                    if shape.Type == 17:
                        shape.TextEffect.Text = shape_data['text']
                    shape.Copy()
                    # print("\nori.name:", shape.Name)
                    # 等待剪贴板数据到来
                    time.sleep(0.5)
                    target_sheet.Paste()

                # 将形状摆到正确的位置
                for idx, (ori, new) in enumerate(zip(sheet_data['shape_info'], target_sheet.Shapes)):
                    new.Left = ori['left']
                    new.Top = ori['top']
                    new.Width = ori['width']
                    new.Height = ori['height']
                    print("\nnew.name:", new.Name)
                    # for shape in target_sheet.Shapes:
                    #     print(f"ID:{shape.ID}, name:{shape.Name}, type:{shape.Type}")

        # 方法2：从已经保存png文件添加到excel中
        elif method == 'draw':
            for sheet_data in data:
                sheet = output_workbook.Worksheets(sheet_data['sheet'])
                for shape_data in sheet_data['shape_info']['shape']:
                    img_path = f"{sheet_data['shape_info']['shape_path']}/{sheet_data['sheet']}_{shape_data['name']}.png"
                    img = sheet.Pictures().Insert(img_path)
                    img.Left = shape_data['left']
                    img.Top = shape_data['top']
                    img.Width = shape_data['width']
                    img.Height = shape_data['height']

        # 存储文件
        output_workbook.SaveAs(output_file)
        print(f"{output_file}已保存成功！")
        return output_file
    except Exception as e:
        print(f"处理形状发生错误：{e}")
    finally:
        excelApp.Quit()
        # 清除缓存图片
        shutil.rmtree(TEMP_PATH)


def apply_excel_template(data, template=None):
    """根据模板更新内容"""
    if template is None:
        return data


def xls_to_xlsx(input_file, output_file=None):
    """
    将xls文件转换为xlsx格式
    """
    # 如果未指定输出路径，自动生成
    if output_file is None:
        output_file = input_file.replace('.xls', '.xlsx')

    # 启动Word应用程序
    excel = wc.Dispatch('Excel.Application')
    excel.Visible = False  # 不显示Excel界面

    try:
        # 打开xls文档
        workbook = excel.Workbooks.Open(input_file)
        # 另存为xlsx格式
        workbook.SaveAs(output_file, FileFormat=16)  # 16表示xlsx格式
        workbook.Close()
        logger.info(f"转换成功: {input_file} -> {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"转换失败: {e}")
        return False
    finally:
        excel.Quit()


def ensure_excel_closed():
    """
    确保所有Excel实例都已关闭
    """
    try:
        # 尝试通过COM获取Excel应用程序
        excel = wc.Dispatch("Excel.Application")
        # 关闭所有工作簿
        for workbook in excel.Workbooks:
            workbook.Close(SaveChanges=False)
        # 退出Excel
        excel.Quit()
        logger.info("已强制关闭所有Excel实例")
    except:
        # 如果无法通过COM关闭，尝试通过任务管理器关闭
        try:
            os.system('taskkill /f /im excel.exe')
            logger.warning("已通过强制方式关闭Excel进程")
        except:
            logger.error("无法关闭Excel进程")


if __name__ == "__main__":
    # 获取当前时间戳
    current_time = datetime.datetime.now().strftime('%y%m%d%H%M%S')
    language = ['英语', '越南语']
    # language = ['英语']

    input_file = r"D:\Code\Project\tools\data\test\5.xlsx"
    output_folder = r"D:\Code\Project\tools\data\temp"

    translator = Translator()
    print(f"********************start at {current_time}********************")

    # 确保没有残留的Excel进程
    ensure_excel_closed()
    # 等待一段时间确保Excel完全关闭
    time.sleep(2)

    # 1. 读取原始内容
    content_data = get_content(input_file)

    # 2. 翻译
    translated_data = add_translation(content_data, translator, language, 'replace_multi')

    # 3. 处理数据内容，使其符合创建新文档格式
    formatted_data = apply_excel_template(translated_data)

    # 4. 创建新文档，替换单元格内容
    create_new_excel(output_folder, formatted_data, 'simple', input_file)
