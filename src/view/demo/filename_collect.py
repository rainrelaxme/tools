#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
@Project: vector_distance.py
@Version:
@Author: RainRelaxMe
@Date: 2025/3/10 21:36
"""
"""
用来将文件夹下的所有文件名，整理输出到一个Excel下。
"""

import os
from openpyxl import Workbook
from moviepy.editor import VideoFileClip
from datetime import timedelta


def get_video_duration(video_path):
    """获取视频时长（秒），失败返回 None"""
    try:
        clip = VideoFileClip(video_path)
        duration = clip.duration
        clip.close()
        return duration
    except Exception as e:
        print(f"无法获取视频时长 {video_path}: {e}")
        return None


def format_duration(seconds):
    """将秒数格式化为 HH:MM:SS"""
    if seconds is None:
        return "N/A"
    return str(timedelta(seconds=seconds))


def extract_file_info_to_excel(folder_path, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "File Info"

    # 表头
    ws['A1'] = "文件路径"
    ws['B1'] = "文件名"
    ws['C1'] = "文件大小 (MB)"
    ws['D1'] = "视频时长 (HH:MM:SS)"
    ws['E1'] = "文件类型"

    row_idx = 2

    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)  # 转为MB
            file_ext = os.path.splitext(filename)[1].lower()

            # 如果是视频文件，尝试获取时长
            video_duration = None
            if file_ext in ('.mp4', '.avi', '.mov', '.mkv', '.flv'):
                video_duration = get_video_duration(file_path)

            # 写入Excel
            ws[f'A{row_idx}'] = file_path
            ws[f'B{row_idx}'] = filename
            ws[f'C{row_idx}'] = round(file_size_mb, 0)  # 保留2位小数
            ws[f'D{row_idx}'] = format_duration(video_duration)
            ws[f'E{row_idx}'] = file_ext

            row_idx += 1

    wb.save(output_excel_path)
    print(f"文件信息已保存到 {output_excel_path}，共找到 {row_idx - 2} 个文件")


# 示例用法
folder_path = r"H:\移动"  # 使用原始字符串避免转义问题
output_excel_path = "file_info_with_duration.xlsx"
extract_file_info_to_excel(folder_path, output_excel_path)
