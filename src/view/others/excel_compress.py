#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
@Project: tools
@File   : excel_compress.py
@Version:
@Author : RainRelaxMe
@Date   : 2025/9/2 0:32
@Info   : 压缩Excel工作表中图片的大小
"""

import os
import zipfile
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import tempfile
import shutil
import time
from datetime import datetime


def compress_excel_images(input_file, output_file, sheet_names, quality=50, max_width=1200, max_height=800,
                          force_convert_jpeg=True):
    """
    压缩Excel工作表中图片的函数

    参数:
    input_file: 输入的Excel文件路径
    output_file: 输出的Excel文件路径
    sheet_names: 要处理的工作表名称列表
    quality: 图片质量 (1-100)，默认50（更激进）
    max_width: 图片最大宽度，默认1200
    max_height: 图片最大高度，默认800
    force_convert_jpeg: 是否强制转换为JPEG格式，默认True
    """

    try:
        # 加载Excel工作簿
        print("正在加载Excel文件...")
        wb = openpyxl.load_workbook(input_file)

        total_images_processed = 0
        sheets_processed = 0
        total_original_size = 0
        total_compressed_size = 0

        # 处理每个选中的工作表
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                print(f"警告: 工作表 '{sheet_name}' 不存在，跳过")
                continue

            sheet = wb[sheet_name]

            # 获取工作表中的所有图片
            images = []
            for img in sheet._images:
                images.append(img)

            if not images:
                print(f"在工作表 '{sheet_name}' 中没有找到图片")
                continue

            print(f"\n正在处理工作表 '{sheet_name}'，找到 {len(images)} 张图片...")

            # 创建一个临时目录来存储图片
            temp_dir = tempfile.mkdtemp(prefix='excel_img_compress_')

            try:
                # 处理每张图片
                sheet_images_processed = 0
                for i, img in enumerate(images):
                    print(f"  处理图片 {i + 1}/{len(images)}")

                    try:
                        # 获取图片数据（二进制格式）
                        img_data = None
                        if hasattr(img, '_data'):
                            img_data = img._data()
                        elif hasattr(img, 'blip') and hasattr(img.blip, 'rEmbed'):
                            # 对于某些类型的图片
                            try:
                                rId = img.blip.rEmbed
                                img_part = wb.worksheets[wb.index(sheet)]._rels[rId].target_part
                                img_data = img_part.blob
                            except:
                                pass

                        if img_data is None:
                            print(f"  无法获取图片 {i + 1} 的数据，跳过")
                            continue

                        # 记录原始图片大小
                        original_size = len(img_data)
                        total_original_size += original_size

                        # 使用Pillow打开图片
                        try:
                            pil_img = Image.open(BytesIO(img_data))
                        except Exception as e:
                            print(f"  无法打开图片 {i + 1}: {e}，跳过")
                            continue

                        # 保存原始格式
                        original_format = pil_img.format or 'UNKNOWN'

                        # 调整图片尺寸（强制调整到合理大小）
                        original_width, original_height = pil_img.size
                        new_width = original_width
                        new_height = original_height

                        # 强制调整尺寸到合理范围
                        if max_width and original_width > max_width:
                            new_width = max_width
                            new_height = int(original_height * (max_width / original_width))

                        if max_height and new_height > max_height:
                            new_height = max_height
                            new_width = int(new_width * (max_height / new_height))

                        # 如果图片很小，稍微放大压缩效果（Excel中过小的图片可能占用更多空间）
                        if original_width < 300 and original_height < 300:
                            new_width = min(600, original_width * 2)
                            new_height = min(600, original_height * 2)

                        if new_width != original_width or new_height != original_height:
                            pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                            print(f"  调整尺寸: {original_width}x{original_height} -> {new_width}x{new_height}")

                        # 压缩图片 - 使用更激进的设置
                        output_buffer = BytesIO()

                        # 强制转换为JPEG格式（除非需要透明度）
                        if force_convert_jpeg and original_format != 'GIF':
                            if pil_img.mode in ('RGBA', 'LA', 'P'):
                                # 如果图片有透明度，创建白色背景
                                if pil_img.mode == 'RGBA':
                                    background = Image.new('RGB', pil_img.size, (255, 255, 255))
                                    background.paste(pil_img, mask=pil_img.split()[3])
                                    pil_img = background
                                else:
                                    pil_img = pil_img.convert('RGB')

                            # 使用较低的JPEG质量
                            pil_img.save(output_buffer, format='JPEG', optimize=True, quality=max(10, quality))
                            new_format = 'JPEG'
                            print(f"  格式转换: {original_format} -> JPEG (质量: {quality})")

                        else:
                            # 保持原格式但优化
                            if original_format == 'PNG':
                                # 对于PNG，转换为更高效的格式
                                if pil_img.mode in ('RGBA', 'LA'):
                                    pil_img = pil_img.quantize(method=2)
                                else:
                                    pil_img = pil_img.convert('P')
                                pil_img.save(output_buffer, format='PNG', optimize=True)
                            else:
                                pil_img.save(output_buffer, format=original_format, optimize=True)
                            new_format = original_format

                        # 获取压缩后的数据
                        compressed_data = output_buffer.getvalue()
                        compressed_size = len(compressed_data)
                        total_compressed_size += compressed_size

                        size_reduction = (original_size - compressed_size) / original_size * 100
                        print(
                            f"  大小: {original_size / 1024:.1f}KB -> {compressed_size / 1024:.1f}KB (-{size_reduction:.1f}%)")

                        # 直接使用内存中的图片数据创建新的图片对象
                        try:
                            # 创建临时内存文件
                            temp_buffer = BytesIO(compressed_data)
                            new_img = XLImage(temp_buffer)

                            # 复制锚点信息
                            if hasattr(img, 'anchor'):
                                new_img.anchor = img.anchor

                            # 替换图片
                            sheet._images.remove(img)
                            sheet.add_image(new_img)

                            sheet_images_processed += 1
                            total_images_processed += 1

                        except Exception as e:
                            print(f"  创建新图片对象时出错: {e}，跳过此图片")
                            continue

                    except Exception as e:
                        print(f"  处理图片 {i + 1} 时出错: {e}，跳过")
                        continue

                print(f"  工作表 '{sheet_name}' 处理完成，成功压缩 {sheet_images_processed}/{len(images)} 张图片")
                sheets_processed += 1

            finally:
                # 清理临时目录
                try:
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)
                except Exception as e:
                    print(f"  清理临时目录时出错: {e}")

        if sheets_processed == 0:
            print("没有找到任何包含图片的工作表")
            return False

        # 保存工作簿
        print(f"\n正在保存工作簿...")
        wb.save(output_file)

        # 显示压缩统计信息
        print(f"\n处理完成!")
        print(f"总共在 {sheets_processed} 个工作表中压缩了 {total_images_processed} 张图片")
        print(f"图片总大小: {total_original_size / 1024 / 1024:.2f}MB -> {total_compressed_size / 1024 / 1024:.2f}MB")

        if total_original_size > 0:
            total_reduction = (total_original_size - total_compressed_size) / total_original_size * 100
            print(f"图片压缩率: -{total_reduction:.1f}%")

        print(f"输出文件: {output_file}")
        return True

    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def get_excel_sheet_names(file_path):
    """获取Excel文件中的所有工作表名称"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"无法读取Excel文件: {e}")
        return []


def select_multiple_sheets(available_sheets):
    """让用户选择多个工作表"""
    print("\n可用的工作表:")
    for i, name in enumerate(available_sheets, 1):
        print(f"{i}. {name}")

    print("\n选择方式:")
    print("1. 输入工作表编号（多个用逗号分隔，如：1,3,5）")
    print("2. 输入工作表名称（多个用逗号分隔）")
    print("3. 输入 'all' 选择所有工作表")
    print("4. 直接回车跳过不选择")

    selection = input("\n请选择要处理的工作表: ").strip()

    if not selection:
        return []

    if selection.lower() == 'all':
        return available_sheets.copy()

    selected_sheets = []

    # 尝试按编号选择
    if all(part.strip().isdigit() for part in selection.split(',')):
        indices = [int(part.strip()) - 1 for part in selection.split(',')]
        for index in indices:
            if 0 <= index < len(available_sheets):
                selected_sheets.append(available_sheets[index])
            else:
                print(f"警告: 编号 {index + 1} 无效，跳过")

    # 尝试按名称选择
    elif any(part.strip() in available_sheets for part in selection.split(',')):
        for part in selection.split(','):
            sheet_name = part.strip()
            if sheet_name in available_sheets:
                selected_sheets.append(sheet_name)
            else:
                print(f"警告: 工作表 '{sheet_name}' 不存在，跳过")

    else:
        print("输入格式无效，请重新选择")
        return select_multiple_sheets(available_sheets)

    # 去重
    selected_sheets = list(dict.fromkeys(selected_sheets))

    if not selected_sheets:
        print("没有选择任何有效的工作表")
        return []

    print(f"\n已选择的工作表: {', '.join(selected_sheets)}")
    return selected_sheets


def add_timestamp_to_filename(file_path):
    """在文件名中添加时间戳"""
    base, ext = os.path.splitext(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_compressed_{timestamp}{ext}"


def main():
    try:
        # 输入文件路径
        input_file = input("请输入Excel文件路径: ").strip().strip('"')

        if not os.path.exists(input_file):
            print("文件不存在!")
            return

        # 获取所有工作表名称
        print("正在读取工作表信息...")
        sheet_names = get_excel_sheet_names(input_file)
        if not sheet_names:
            return

        # 选择要处理的工作表
        selected_sheets = select_multiple_sheets(sheet_names)
        if not selected_sheets:
            print("没有选择任何工作表，程序退出")
            return

        # 图片质量（使用更低的默认值）
        quality_input = input("请输入图片质量 (1-100, 默认50，建议20-70): ").strip()
        try:
            quality = int(quality_input) if quality_input else 50
            quality = max(1, min(100, quality))
        except ValueError:
            print("输入无效，使用默认质量50")
            quality = 50

        # 最大尺寸
        max_width_input = input("请输入图片最大宽度 (默认1200): ").strip()
        max_width = int(max_width_input) if max_width_input and max_width_input.isdigit() else 1200

        max_height_input = input("请输入图片最大高度 (默认800): ").strip()
        max_height = int(max_height_input) if max_height_input and max_height_input.isdigit() else 800

        # 是否强制转换为JPEG
        convert_jpeg = input("是否强制转换为JPEG格式以提高压缩率? (y/n, 默认y): ").strip().lower()
        force_convert_jpeg = convert_jpeg != 'n'

        # 输出文件路径（自动添加时间戳）
        default_output = add_timestamp_to_filename(input_file)
        output_file_input = input(
            f"请输入输出文件路径 (直接回车使用默认名称 '{os.path.basename(default_output)}'): ").strip().strip('"')

        if output_file_input:
            output_file = output_file_input
        else:
            output_file = default_output

        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 执行压缩
        success = compress_excel_images(
            input_file, output_file, selected_sheets,
            quality=quality, max_width=max_width, max_height=max_height,
            force_convert_jpeg=force_convert_jpeg
        )

        if success:
            # 比较整个文件的大小
            try:
                input_size = os.path.getsize(input_file)
                output_size = os.path.getsize(output_file)
                reduction = (input_size - output_size) / input_size * 100 if input_size > 0 else 0

                print(f"\n最终文件压缩结果:")
                print(f"原始文件大小: {input_size / 1024 / 1024:.2f} MB")
                print(f"压缩后文件大小: {output_size / 1024 / 1024:.2f} MB")
                print(f"总体减小了: {reduction:.2f}%")

                if reduction < 5:
                    print("\n提示: 压缩效果不明显，可能是因为:")
                    print("1. 图片不是文件大小的主要部分")
                    print("2. Excel文件中有其他大量数据")
                    print("3. 图片原本就已经被压缩过")
                    print("建议尝试更低的图片质量设置(20-40)")

            except:
                print("无法比较文件大小，但处理已完成")

    except KeyboardInterrupt:
        print("\n用户中断操作")
    except Exception as e:
        print(f"程序运行出错: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    print("Excel图片压缩工具（增强版）")
    print("=" * 60)
    print("提示: 为了获得更好的压缩效果，建议:")
    print("- 使用较低的图片质量 (20-50)")
    print("- 启用强制JPEG转换")
    print("- 设置合理的最大尺寸")
    print("=" * 60)
    main()